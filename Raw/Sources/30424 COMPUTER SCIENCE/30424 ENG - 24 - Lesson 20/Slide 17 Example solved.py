
# To avoid having to specify the Path,
# make sure that both files .py and .xlsx
# are saved in the same folder

import openpyxl

MyWorkbook = openpyxl.load_workbook('companies.xlsx')
MyWorksheet = MyWorkbook['Feb']
idx = 7
companies_list = []

while MyWorksheet.cell(row=idx, column=3).value != None:
    companies_list.append(MyWorksheet.cell(row=idx, column=3).value)
    idx = idx + 1

print('\nCompanies:\n')
for company in companies_list:
    print(company)
