
# To avoid having to specify the Path,
# make sure that both files .py and .xlsx
# are saved in the same folder

import openpyxl

WBname = 'Lesson20.xlsx'
WSname = 'Invoices'


myWB = openpyxl.load_workbook(WBname)
myWS = myWB[WSname]


MaxR = myWS.max_row
MaxC = myWS.max_column


KeysList = []
for col in range(1, MaxC + 1):
    KeysList.append(myWS.cell(row=1, column=col).value)

myDct = {}
for k in KeysList:
    myDct[k] = []

myCol = 1
for myList in myDct.values():
    for myRow in range(2, MaxR + 1):
        myList.append(myWS.cell(row=myRow, column=myCol).value)
    myCol = myCol + 1



print("Filled dictionary:", myDct)





